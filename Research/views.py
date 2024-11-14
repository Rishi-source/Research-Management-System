from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from .forms import *
from .models import *
from django.urls import reverse
from django.contrib import messages
from datetime import datetime
import matplotlib.pyplot as plt
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from django.http import HttpResponse
from datetime import datetime
from openpyxl.utils.cell import get_column_letter
import json

plt.switch_backend("Agg")


def login_user(request):
    if "error" in request.session:
        del request.session["error"]

    username = request.POST["username"]
    password = request.POST["password"]

    user = authenticate(username=username, password=password)

    if user != None:
        login(request, user)
        request.session["username"] = username
        return redirect("dashboard")
    else:
        request.session["error"] = "Username or Password is incorrect"
        return redirect("login")


def loginView(request):
    if "username" in request.session:
        return redirect("dashboard")
    elif request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = request.POST["username"]
            password = request.POST["password"]
            user = authenticate(username=username, password=password)
            if user != None:
                login(request, user)
                request.session["username"] = username
                return redirect("dashboard")
            else:
                request.session["error"] = "Username or Password is incorrect"
                return redirect("login")

    else:
        form = LoginForm()

    return render(request, template_name="login/login.html", context={"form": form})


def logout_user(request):
    logout(request)
    if "username" in request.session:
        del request.session["username"]
    return redirect("login")


def dashboard(request):
    if not request.user.is_authenticated:
        return redirect("login")

    ongoing_projects_count = EndorsementForm.objects.filter(status="ongoing").count()
    completed_projects_count = EndorsementForm.objects.filter(
        status="completed"
    ).count()

    ongoing_sanctioned_total = (
        Budget.objects.filter(project__status="ongoing").aggregate(
            total_sanctioned=Sum(
                ExpressionWrapper(
                    F("Equipment")
                    + F("Consumables")
                    + F("Fellowship")
                    + F("Contingency")
                    + F("Travel")
                    + F("Field_Testing")
                    + F("Miscellaneous"),
                    output_field=DecimalField(),
                )
            )
        )["total_sanctioned"]
        or 0
    )

    completed_sanctioned_total = (
        Budget.objects.filter(project__status="completed").aggregate(
            total_sanctioned=Sum(
                ExpressionWrapper(
                    F("Equipment")
                    + F("Consumables")
                    + F("Fellowship")
                    + F("Contingency")
                    + F("Travel")
                    + F("Field_Testing")
                    + F("Miscellaneous"),
                    output_field=DecimalField(),
                )
            )
        )["total_sanctioned"]
        or 0
    )

    context = {
        "ongoing_projects_count": ongoing_projects_count,
        "completed_projects_count": completed_projects_count,
        "ongoing_sanctioned_total": ongoing_sanctioned_total,
        "completed_sanctioned_total": completed_sanctioned_total,
    }

    departments = Department.objects.all()
    department_data = {"departments": [], "amounts": []}

    for dept in departments:
        total_sanctioned = (
            EndorsementForm.objects.filter(department=dept).aggregate(
                total=Sum(F("CAPEX") + F("OPEX"))
            )["total"]
            or 0
        )

        if total_sanctioned > 0:
            department_data["departments"].append(dept.name)
            department_data["amounts"].append(float(total_sanctioned))

    context.update({"department_data": json.dumps(department_data)})

    return render(request, "management/dashboard.html", context)


def add_pi_view(request):
    if not request.user.is_authenticated:
        return redirect("login")

    if request.method == "POST":
        PSRN = request.POST.get("PSRN")
        Name_PI = request.POST.get("Name_PI")
        Designation_PI = request.POST.get("Designation_PI")
        organization_id = request.POST.get("organization")
        department_id = request.POST.get("department")
        superannuation_date = request.POST.get("superannuation_date")

        organization = Organization.objects.get(id=organization_id)
        department = Department.objects.get(id=department_id)

        new_pi = PrincipleInvestigator(
            user=request.user,
            PSRN=PSRN,
            Name_PI=Name_PI,
            Designation_PI=Designation_PI,
            organization=organization,
            department=department,
            superannuation_date=superannuation_date,
        )
        new_pi.save()

        return redirect(reverse("add_project") + "?data_saved=true")

    organizations = Organization.objects.all()
    departments = Department.objects.all()

    return render(
        request,
        "management/add_pi.html",
        {"organizations": organizations, "departments": departments},
    )


def project_list(request):
    if not request.user.is_authenticated:
        return redirect("login")
    projects = EndorsementForm.objects.all()
    context = {"projects": projects}
    return render(request, "management/add_details.html", context)


def add_budget_view(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)

    budget = Budget.objects.filter(project=project).first()

    if request.method == "POST":
        if not budget:

            budget = Budget(project=project)

        sanctiondate = request.POST.get("sanctiondate")
        sanctionorder = request.POST.get("sanctionorder")

        if sanctiondate:
            budget.sanctiondate = sanctiondate
        if sanctionorder:
            budget.sanctionorder = sanctionorder

        budget.Equipment = request.POST.get("Equipment", 0)
        budget.Consumables = request.POST.get("Consumables", 0)
        budget.Fellowship = request.POST.get("Fellowship", 0)
        budget.Contingency = request.POST.get("Contingency", 0)
        budget.Travel = request.POST.get("Travel", 0)
        budget.Field_Testing = request.POST.get("Field_Testing", 0)
        budget.Miscellaneous = request.POST.get("Miscellaneous", 0)

        try:
            budget.save()
            messages.success(request, "Budget details saved successfully!")
        except Exception as e:
            messages.error(request, f"An error occurred while saving the budget: {e}")

        return redirect("project_list")

    return render(
        request,
        "management/add_sanctioned_amount.html",
        {"project": project, "budget": budget},
    )


def add_received_amount_view(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)

    received_amount = RecievedAmount.objects.filter(project=project).first()

    if request.method == "POST":
        if not received_amount:

            received_amount = RecievedAmount(project=project)

        received_amount.Equipment = request.POST.get("Equipment", 0)
        received_amount.Consumables = request.POST.get("Consumables", 0)
        received_amount.Fellowship = request.POST.get("Fellowship", 0)
        received_amount.Contingency = request.POST.get("Contingency", 0)
        received_amount.Travel = request.POST.get("Travel", 0)
        received_amount.Field_Testing = request.POST.get("Field_Testing", 0)
        received_amount.Miscellaneous = request.POST.get("Miscellaneous", 0)

        try:
            received_amount.save()
            messages.success(request, "Received amount details saved successfully!")
        except Exception as e:
            messages.error(
                request, f"An error occurred while saving the received amount: {e}"
            )

        return redirect("project_list")

    return render(
        request,
        "management/add_received_amount.html",
        {"project": project, "received_amount": received_amount},
    )


def add_installment(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)
    installments = Installment.objects.filter(project=project)

    if request.method == "POST":
        installment_year_id = request.POST.get("installment_year")
        amount = request.POST.get("amount")
        file = request.FILES.get("file")
        if file:
            print(file)
        else:
            print("No file")

        if installments.filter(installment_year_id=installment_year_id).exists():
            messages.error(request, "An entry for this financial year already exists.")
        else:
            financial_year = get_object_or_404(FinancialYear, id=installment_year_id)
            Installment.objects.create(
                project=project,
                installment_year=financial_year,
                amount=amount,
                file=file,
            )
            messages.success(request, "Installment added successfully.")
            return redirect("add_installment", project_id=project.id)

    financial_years = FinancialYear.objects.exclude(
        id__in=installments.values_list("installment_year_id", flat=True)
    )
    context = {
        "project": project,
        "installments": installments,
        "financial_years": financial_years,
    }
    return render(request, "management/add_installment.html", context)


def edit_installment(request, project_id, installment_id):
    if not request.user.is_authenticated:
        return redirect("login")
    installment = get_object_or_404(
        Installment, id=installment_id, project_id=project_id
    )

    if request.method == "POST":
        amount = request.POST.get("amount")
        installment.amount = amount
        file = request.FILES.get("file")
        if file:
            installment.file = file

        installment.save()
        messages.success(request, "Installment updated successfully.")
        return redirect("add_installment", project_id=project_id)

    context = {
        "installment": installment,
        "project": installment.project,
    }
    return render(request, "management/edit_installment.html", context)


def delete_installment(request, project_id, installment_id):
    if not request.user.is_authenticated:
        return redirect("login")
    installment = get_object_or_404(
        Installment, id=installment_id, project_id=project_id
    )
    installment.delete()
    messages.success(request, "Installment deleted successfully.")
    return redirect("add_installment", project_id=project_id)


def add_expenditure(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)
    if request.method == "POST":
        year_id = request.POST.get("year_expenditure")
        expenditure_type = request.POST.get("expenditure_type")
        expenditure_date = request.POST.get("Expenditure_date")
        amount = request.POST.get("amount")
        type = request.POST.get("type", "")

        try:
            expenditure_date = datetime.strptime(expenditure_date, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect("add_expenditure", project_id=project_id)

        year = get_object_or_404(FinancialYear, id=year_id)

        Expenditure.objects.create(
            year=year,
            expenditure_type=expenditure_type,
            Expenditure_date=expenditure_date,
            amount=amount,
            type=type,
            project=project,
        )
        messages.success(request, "Expenditure added successfully.")
        return redirect("add_expenditure", project_id=project_id)

    expenditures = Expenditure.objects.filter(project=project)
    financial_years = FinancialYear.objects.all()
    context = {
        "project": project,
        "expenditures": expenditures,
        "financial_years": financial_years,
    }
    return render(request, "management/add_expenditure.html", context)


def edit_expenditure(request, expenditure_id):
    if not request.user.is_authenticated:
        return redirect("login")
    expenditure = get_object_or_404(Expenditure, id=expenditure_id)
    financial_years = FinancialYear.objects.all()
    expenditure_types = [
        "Equipment",
        "Consumables",
        "Fellowship",
        "Contingency",
        "Travel",
        "Field Testing",
        "Miscellaneous",
    ]

    if request.method == "POST":
        year_id = request.POST.get("year_expenditure")
        expenditure_type = request.POST.get("expenditure_type")
        expenditure_date = request.POST.get("Expenditure_date")
        amount = request.POST.get("amount")
        type = request.POST.get("type")

        try:
            expenditure_date = datetime.strptime(expenditure_date, "%Y-%m-%d").date()
            financial_year = get_object_or_404(FinancialYear, id=year_id)
            if not amount.isdigit():
                raise ValueError("Amount must be a number.")

            expenditure.year = financial_year
            expenditure.expenditure_type = expenditure_type
            expenditure.Expenditure_date = expenditure_date
            expenditure.amount = int(amount)
            expenditure.type = type
            expenditure.save()

            messages.success(request, "Expenditure updated successfully.")
            return redirect("project_list")
        except ValueError as e:
            messages.error(request, str(e))
            return redirect("edit_expenditure", expenditure_id=expenditure_id)

    context = {
        "expenditure": expenditure,
        "financial_years": financial_years,
        "expenditure_types": expenditure_types,
    }
    return render(request, "management/edit_expenditure.html", context)


def delete_expenditure(request, expenditure_id):
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method == "POST":
        expenditure = get_object_or_404(Expenditure, id=expenditure_id)
        project_id = expenditure.project.id
        expenditure.delete()
        messages.success(request, "Expenditure deleted successfully.")
        return redirect("add_expenditure", project_id=project_id)
    else:

        messages.error(request, "Invalid request method.")
        return redirect("add_expenditure", project_id=project_id)


def edit_project(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    endorsement = get_object_or_404(EndorsementForm, id=project_id)

    if request.method == "POST":

        endorsement.Project_Title = request.POST.get(
            "Project_Title", endorsement.Project_Title
        )
        endorsement.PSRN = request.POST.get("PSRN", endorsement.PSRN)
        endorsement.Name_PI = request.POST.get("Name_PI", endorsement.Name_PI)
        endorsement.Designation_PI = request.POST.get(
            "Designation_PI", endorsement.Designation_PI
        )

        organization_id = request.POST.get("organization")
        department_id = request.POST.get("department")
        financial_year_id = request.POST.get("year_project_sanctioned")

        if organization_id:
            try:
                endorsement.organization = Organization.objects.get(id=organization_id)
            except Organization.DoesNotExist:
                messages.error(request, "Invalid organization selected.")
                return redirect("edit_project", project_id=project_id)

        if department_id:
            try:
                endorsement.department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                messages.error(request, "Invalid department selected.")
                return redirect("edit_project", project_id=project_id)

        if financial_year_id:
            try:
                endorsement.year = FinancialYear.objects.get(id=financial_year_id)
            except FinancialYear.DoesNotExist:
                messages.error(request, "Invalid financial year selected.")
                return redirect("edit_project", project_id=project_id)

        endorsement.project_type = request.POST.get(
            "project_type", endorsement.project_type
        )
        endorsement.funding_agency = request.POST.get(
            "funding_agency", endorsement.funding_agency
        )
        endorsement.funding_scheme = request.POST.get(
            "funding_scheme", endorsement.funding_scheme
        )
        endorsement.funding_agency_type = request.POST.get(
            "funding_agency_type", endorsement.funding_agency_type
        )
        endorsement.CAPEX = request.POST.get("CAPEX", endorsement.CAPEX)
        endorsement.OPEX = request.POST.get("OPEX", endorsement.OPEX)
        endorsement.proposal_file = request.POST.get(
            "project_proposal", endorsement.proposal_file
        )
        endorsement.project_duration = request.POST.get(
            "project_duration", endorsement.project_duration
        )
        endorsement.sanction_letter = request.POST.get(
            "sanction_letter", endorsement.sanction_letter
        )

        commencement_date = request.POST.get("commencement_date_of_project")
        endorsement.commencement_date_of_project = (
            commencement_date if commencement_date else None
        )

        end_date = request.POST.get("end_date_of_project")
        endorsement.end_date = end_date if end_date else None

        endorsement.save()

        co_pi_ids = request.POST.getlist("pi_select")
        co_pis = PrincipleInvestigator.objects.filter(id__in=co_pi_ids)
        endorsement.Co_PI1 = co_pis[0] if len(co_pis) > 0 else None
        endorsement.Co_PI2 = co_pis[1] if len(co_pis) > 1 else None
        endorsement.Co_PI3 = co_pis[2] if len(co_pis) > 2 else None
        endorsement.Co_PI4 = co_pis[3] if len(co_pis) > 3 else None
        endorsement.Co_PI5 = co_pis[4] if len(co_pis) > 4 else None
        endorsement.Co_PI6 = co_pis[5] if len(co_pis) > 5 else None
        endorsement.save()

        messages.success(request, "Project updated successfully!")
        return redirect("project_list")

    organizations = Organization.objects.all()
    departments = Department.objects.all()
    principal_investigators = PrincipleInvestigator.objects.filter(user=request.user)
    financial_years = FinancialYear.objects.all()
    co_pi_ids = [
        endorsement.Co_PI1_id,
        endorsement.Co_PI2_id,
        endorsement.Co_PI3_id,
        endorsement.Co_PI4_id,
        endorsement.Co_PI5_id,
        endorsement.Co_PI6_id,
    ]

    context = {
        "endorsement": endorsement,
        "organizations": organizations,
        "departments": departments,
        "principal_investigators": principal_investigators,
        "financial_years": financial_years,
        "co_pi_ids": co_pi_ids,
    }

    return render(request, "management/edit_project.html", context)


def delete_project(request):
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method == "POST":
        project_id = request.POST.get("project_id")
        project = get_object_or_404(EndorsementForm, id=project_id)
        project.delete()
        messages.success(request, "Project successfully deleted.")
        return redirect("project_list")


def view_projects(request):
    if not request.user.is_authenticated:
        return redirect("login")
    projects = EndorsementForm.objects.all()

    project_data = []

    for project in projects:
        budget = Budget.objects.filter(project=project).first()
        received_amount = RecievedAmount.objects.filter(project=project).first()
        installments = Installment.objects.filter(project=project)
        expenditures = Expenditure.objects.filter(project=project)

        total_sanctioned = 0
        total_received = 0
        total_installments = installments.count()
        total_expenditure = 0
        sanctionorder = "No Sanction Order Available"
        sanctiondate = "No Sanction Date Available"

        if budget:
            total_sanctioned = (
                budget.Equipment
                + budget.Consumables
                + budget.Fellowship
                + budget.Contingency
                + budget.Travel
                + budget.Field_Testing
                + budget.Miscellaneous
            )

            if budget.sanctionorder:
                sanctionorder = budget.sanctionorder

            if hasattr(budget, "sanctiondate"):
                sanctiondate = budget.sanctiondate

        if received_amount:
            total_received = (
                received_amount.Equipment
                + received_amount.Consumables
                + received_amount.Fellowship
                + received_amount.Contingency
                + received_amount.Travel
                + received_amount.Field_Testing
                + received_amount.Miscellaneous
            )

        if expenditures:
            total_expenditure = sum(exp.amount for exp in expenditures)

        project_data.append(
            {
                "project": project,
                "total_sanctioned": total_sanctioned,
                "total_received": total_received,
                "total_installments": total_installments,
                "total_expenditure": total_expenditure,
                "sanctionorder": sanctionorder,
                "sanctiondate": sanctiondate,
            }
        )
    departments = Department.objects.all()
    financialyear = FinancialYear.objects.all()
    context = {
        "projects": project_data,
        "departments": departments,
        "years": financialyear,
    }
    return render(request, "management/view_projects.html", context)


def mark_completed(request, pk):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, pk=pk)
    project.status = "completed"
    project.save()
    messages.success(request, f"Project {project.Project_Title} marked as completed.")
    return redirect("view_projects")


def edit_dates(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, pk=project_id)

    if request.method == "POST":
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        if start_date:
            project.start_date = start_date
        if end_date:
            project.end_date = end_date

        project.save()
        return redirect("view_projects")

    context = {
        "project": project,
    }
    return render(request, "management/edit_dates.html", context)


def project_details(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)
    principal_investigators = PrincipleInvestigator.objects.filter(
        id__in=[
            project.Co_PI1_id,
            project.Co_PI2_id,
            project.Co_PI3_id,
            project.Co_PI4_id,
            project.Co_PI5_id,
            project.Co_PI6_id,
        ]
    ).exclude(id=None)

    budget = Budget.objects.filter(project=project).first()
    received_amount = RecievedAmount.objects.filter(project=project).first()
    installments = Installment.objects.filter(project=project)
    expenditures = Expenditure.objects.filter(project=project)
    fin_year_budgets = FinYearBudget.objects.filter(project=project)

    types = [
        "Equipment",
        "Consumables",
        "Fellowship",
        "Contingency",
        "Travel",
        "Field Testing",
        "Miscellaneous",
    ]
    total_by_type = {
        exp_type: expenditures.filter(expenditure_type=exp_type).aggregate(
            Sum("amount")
        )["amount__sum"]
        or 0
        for exp_type in types
    }

    total_sanctioned = (
        budget.Equipment
        + budget.Consumables
        + budget.Fellowship
        + budget.Contingency
        + budget.Travel
        + budget.Field_Testing
        + budget.Miscellaneous
        if budget
        else 0
    )
    total_received = (
        received_amount.Equipment
        + received_amount.Consumables
        + received_amount.Fellowship
        + received_amount.Contingency
        + received_amount.Travel
        + received_amount.Field_Testing
        + received_amount.Miscellaneous
        if received_amount
        else 0
    )
    total_installments_amount = (
        installments.aggregate(Sum("amount"))["amount__sum"] or 0
    )
    total_expenditure = expenditures.aggregate(Sum("amount"))["amount__sum"] or 0

    balance_equipment_sanctioned = (budget.Equipment if budget else 0) - (
        expenditures.filter(expenditure_type="Equipment").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_consumables_sanctioned = (budget.Consumables if budget else 0) - (
        expenditures.filter(expenditure_type="Consumables").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_fellowship_sanctioned = (budget.Fellowship if budget else 0) - (
        expenditures.filter(expenditure_type="Fellowship").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_contingency_sanctioned = (budget.Contingency if budget else 0) - (
        expenditures.filter(expenditure_type="Contingency").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_travel_sanctioned = (budget.Travel if budget else 0) - (
        expenditures.filter(expenditure_type="Travel").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_field_testing_sanctioned = (budget.Field_Testing if budget else 0) - (
        expenditures.filter(expenditure_type="Field Testing").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_miscellaneous_sanctioned = (budget.Miscellaneous if budget else 0) - (
        expenditures.filter(expenditure_type="Miscellaneous").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )

    balance_equipment_received = (
        received_amount.Equipment if received_amount else 0
    ) - (
        expenditures.filter(expenditure_type="Equipment").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_consumables_received = (
        received_amount.Consumables if received_amount else 0
    ) - (
        expenditures.filter(expenditure_type="Consumables").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_fellowship_received = (
        received_amount.Fellowship if received_amount else 0
    ) - (
        expenditures.filter(expenditure_type="Fellowship").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_contingency_received = (
        received_amount.Contingency if received_amount else 0
    ) - (
        expenditures.filter(expenditure_type="Contingency").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_travel_received = (received_amount.Travel if received_amount else 0) - (
        expenditures.filter(expenditure_type="Travel").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_field_testing_received = (
        received_amount.Field_Testing if received_amount else 0
    ) - (
        expenditures.filter(expenditure_type="Field Testing").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )
    balance_miscellaneous_received = (
        received_amount.Miscellaneous if received_amount else 0
    ) - (
        expenditures.filter(expenditure_type="Miscellaneous").aggregate(Sum("amount"))[
            "amount__sum"
        ]
        or 0
    )

    balance_sanctioned = total_sanctioned - total_expenditure
    balance_received = total_received - total_expenditure
    expenditures = Expenditure.objects.filter(project=project)

    financial_year_expenditure_totals = {}

    expenditures = Expenditure.objects.filter(project=project)

    types = [
        "Equipment",
        "Consumables",
        "Fellowship",
        "Contingency",
        "Travel",
        "Field Testing",
        "Miscellaneous",
    ]
    financial_year_expenditure_totals = {}

    for expenditure in expenditures:
        year_key = expenditure.year.name if expenditure.year else "Unknown Year"
        if year_key not in financial_year_expenditure_totals:
            financial_year_expenditure_totals[year_key] = {type_: 0 for type_ in types}
        exp_type = expenditure.expenditure_type
        financial_year_expenditure_totals[year_key][exp_type] += expenditure.amount

    bifurcation_data = []
    for fy_budget in fin_year_budgets:
        bifurcation_data.append(
            {
                "year": fy_budget.year.name,
                "Equipment": fy_budget.Equipment or 0,
                "Consumables": fy_budget.Consumables or 0,
                "Fellowship": fy_budget.Fellowship or 0,
                "Contingency": fy_budget.Contingency or 0,
                "Travel": fy_budget.Travel or 0,
                "Field_Testing": fy_budget.Field_Testing or 0,
                "Miscellaneous": fy_budget.Miscellaneous or 0,
                "Total": sum(
                    [
                        fy_budget.Equipment or 0,
                        fy_budget.Consumables or 0,
                        fy_budget.Fellowship or 0,
                        fy_budget.Contingency or 0,
                        fy_budget.Travel or 0,
                        fy_budget.Field_Testing or 0,
                        fy_budget.Miscellaneous or 0,
                    ]
                ),
            }
        )
    total_sum_for_all_types = sum(total_by_type.values())

    context = {
        "project": project,
        "principal_investigators": principal_investigators,
        "budget": budget,
        "received_amount": received_amount,
        "installments": installments,
        "expenditures": expenditures,
        "total_sanctioned": total_sanctioned,
        "total_received": total_received,
        "total_installments_amount": total_installments_amount,
        "total_expenditure": total_expenditure,
        "balance_sanctioned": balance_sanctioned,
        "balance_received": balance_received,
        "balance_equipment_sanctioned": balance_equipment_sanctioned,
        "balance_consumables_sanctioned": balance_consumables_sanctioned,
        "balance_fellowship_sanctioned": balance_fellowship_sanctioned,
        "balance_contingency_sanctioned": balance_contingency_sanctioned,
        "balance_travel_sanctioned": balance_travel_sanctioned,
        "balance_field_testing_sanctioned": balance_field_testing_sanctioned,
        "balance_miscellaneous_sanctioned": balance_miscellaneous_sanctioned,
        "balance_equipment_received": balance_equipment_received,
        "balance_consumables_received": balance_consumables_received,
        "balance_fellowship_received": balance_fellowship_received,
        "balance_contingency_received": balance_contingency_received,
        "balance_travel_received": balance_travel_received,
        "balance_field_testing_received": balance_field_testing_received,
        "balance_miscellaneous_received": balance_miscellaneous_received,
        "bifurcation_data": bifurcation_data,
        "total_by_type": total_by_type,
        "financial_year_expenditure_totals": financial_year_expenditure_totals,
        "total_sum_for_all_types": total_sum_for_all_types,
    }

    return render(request, "management/project_details.html", context)


def manage_departments(request):
    if not request.user.is_authenticated:
        return redirect("login")
    departments = Department.objects.all()
    return render(
        request, "management/add_departments.html", {"departments": departments}
    )


def add_department(request):
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method == "POST":
        name = request.POST.get("name")
        if Department.objects.filter(name=name).exists():
            messages.error(request, "Department already exists.")
        else:
            Department.objects.create(name=name)
            messages.success(request, "Department added successfully.")
        return redirect("manage_departments")


def delete_department(request, id):
    if not request.user.is_authenticated:
        return redirect("login")
    department = get_object_or_404(Department, id=id)
    department.delete()
    messages.success(request, "Department deleted successfully.")
    return redirect("manage_departments")


def manage_organizations(request):
    if not request.user.is_authenticated:
        return redirect("login")

    organizations = Organization.objects.all()

    return render(
        request, "management/add_organisations.html", {"organizations": organizations}
    )


def add_organization(request):
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method == "POST":

        name = request.POST.get("name")

        if Organization.objects.filter(name=name).exists():
            messages.error(request, "Organization with this name already exists.")
        else:

            Organization.objects.create(name=name)
            messages.success(request, "Organization added successfully.")

        return redirect("manage_organizations")


def delete_organization(request, id):
    if not request.user.is_authenticated:
        return redirect("login")

    organization = get_object_or_404(Organization, id=id)

    if request.method == "POST":

        organization.delete()
        messages.success(request, "Organization deleted successfully.")

    return redirect("manage_organizations")


def manage_financial_years(request):
    if not request.user.is_authenticated:
        return redirect("login")
    financial_years = FinancialYear.objects.all()
    return render(
        request,
        "management/add_financial_year.html",
        {"financial_years": financial_years},
    )


def add_financial_year(request):
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method == "POST":
        name = request.POST.get("name")
        if FinancialYear.objects.filter(name=name).exists():
            messages.error(request, "Financial year already exists.")
        else:
            FinancialYear.objects.create(name=name)
            messages.success(request, "Financial year added successfully.")
        return redirect("manage_financial_years")


def delete_financial_year(request, id):
    if not request.user.is_authenticated:
        return redirect("login")
    financial_year = get_object_or_404(FinancialYear, id=id)
    financial_year.delete()
    messages.success(request, "Financial year deleted successfully.")
    return redirect("manage_financial_years")


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type="application/pdf")
    return None


def project_details_pdf(request, project_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)
    principal_investigators = PrincipleInvestigator.objects.filter(
        id__in=[
            project.Co_PI1_id,
            project.Co_PI2_id,
            project.Co_PI3_id,
            project.Co_PI4_id,
            project.Co_PI5_id,
            project.Co_PI6_id,
        ]
    ).exclude(id=None)

    budget = Budget.objects.filter(project=project).first()
    received_amount = RecievedAmount.objects.filter(project=project).first()
    installments = Installment.objects.filter(project=project)
    expenditures = Expenditure.objects.filter(project=project)
    fin_year_budgets = FinYearBudget.objects.filter(project=project)

    types = [
        "Equipment",
        "Consumables",
        "Fellowship",
        "Contingency",
        "Travel",
        "Field Testing",
        "Miscellaneous",
    ]
    attribute_mapping = {"Field Testing": "Field_Testing"}

    total_by_type = {
        exp_type: expenditures.filter(expenditure_type=exp_type).aggregate(
            Sum("amount")
        )["amount__sum"]
        or 0
        for exp_type in types
    }

    total_sanctioned = (
        sum(
            [
                getattr(budget, attribute_mapping.get(type_, type_), 0) or 0
                for type_ in types
            ]
        )
        if budget
        else 0
    )

    total_received = (
        sum(
            [
                getattr(received_amount, attribute_mapping.get(type_, type_), 0) or 0
                for type_ in types
            ]
        )
        if received_amount
        else 0
    )

    total_installments_amount = (
        installments.aggregate(Sum("amount"))["amount__sum"] or 0
    )
    total_expenditure = expenditures.aggregate(Sum("amount"))["amount__sum"] or 0

    balance_sanctioned = total_sanctioned - total_expenditure
    balance_received = total_received - total_expenditure

    balance_by_type = {
        f"balance_{type_.lower().replace(' ', '_')}_sanctioned": (
            getattr(budget, attribute_mapping.get(type_, type_)) if budget else 0
        )
        - total_by_type[type_]
        for type_ in types
    }
    balance_by_type.update(
        {
            f"balance_{type_.lower().replace(' ', '_')}_received": (
                getattr(received_amount, attribute_mapping.get(type_, type_))
                if received_amount
                else 0
            )
            - total_by_type[type_]
            for type_ in types
        }
    )

    financial_year_expenditure_totals = {}
    for expenditure in expenditures:
        year_key = expenditure.year.name if expenditure.year else "Unknown Year"
        if year_key not in financial_year_expenditure_totals:
            financial_year_expenditure_totals[year_key] = {type_: 0 for type_ in types}
        exp_type = expenditure.expenditure_type
        financial_year_expenditure_totals[year_key][exp_type] += expenditure.amount

    bifurcation_data = []
    for fy_budget in fin_year_budgets:
        bifurcation_data.append(
            {
                "year": fy_budget.year.name,
                "Equipment": fy_budget.Equipment or 0,
                "Consumables": fy_budget.Consumables or 0,
                "Fellowship": fy_budget.Fellowship or 0,
                "Contingency": fy_budget.Contingency or 0,
                "Travel": fy_budget.Travel or 0,
                "Field_Testing": fy_budget.Field_Testing or 0,
                "Miscellaneous": fy_budget.Miscellaneous or 0,
                "Total": sum(
                    [
                        fy_budget.Equipment or 0,
                        fy_budget.Consumables or 0,
                        fy_budget.Fellowship or 0,
                        fy_budget.Contingency or 0,
                        fy_budget.Travel or 0,
                        fy_budget.Field_Testing or 0,
                        fy_budget.Miscellaneous or 0,
                    ]
                ),
            }
        )

    total_sum_for_all_types = sum(total_by_type.values())

    context = {
        "project": project,
        "principal_investigators": principal_investigators,
        "budget": budget,
        "received_amount": received_amount,
        "installments": installments,
        "expenditures": expenditures,
        "total_sanctioned": total_sanctioned,
        "total_received": total_received,
        "total_installments_amount": total_installments_amount,
        "total_expenditure": total_expenditure,
        "balance_sanctioned": balance_sanctioned,
        "balance_received": balance_received,
        "bifurcation_data": bifurcation_data,
        "total_by_type": total_by_type,
        "financial_year_expenditure_totals": financial_year_expenditure_totals,
        "total_sum_for_all_types": total_sum_for_all_types,
        **balance_by_type,
    }

    pdf = render_to_pdf("management/project_details_pdf.html", context)
    filename = f"Project_Details.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def add_project(request):
    if not request.user.is_authenticated:
        return redirect("login")
    if request.method == "POST":

        project_title = request.POST.get("Project_Title")
        psrn = request.POST.get("PSRN")
        name_pi = request.POST.get("Name_PI")
        designation_pi = request.POST.get("Designation_PI")
        organization_id = request.POST.get("organization")
        department_id = request.POST.get("department")
        project_type = request.POST.get("project_type")
        sanction_letter = request.FILES.get("sanction_letter")
        funding_agency = request.POST.get("funding_agency")
        project_proposal = request.FILES.get("project_proposal")
        funding_agency_type = request.POST.get("funding_agency_type")
        financial_year_id = request.POST.get("year_project_sanctioned")
        capex = request.POST.get("CAPEX")
        opex = request.POST.get("OPEX")
        project_duration = request.POST.get("project_duration")
        co_pi_ids = request.POST.getlist("pi_select")
        start_date = request.POST.get("commencement_date_of_project")
        end_date = request.POST.get("end_date_of_project")

        if not all(
            [
                project_title,
                psrn,
                name_pi,
                designation_pi,
                organization_id,
                department_id,
                project_type,
                sanction_letter,
                project_proposal,
                funding_agency,
                financial_year_id,
                capex,
                opex,
                project_duration,
                start_date,
                end_date,
            ]
        ):
            messages.error(request, "Please fill out all required fields.")
            return redirect("add_project")

        try:
            organization = Organization.objects.get(id=organization_id)
            department = Department.objects.get(id=department_id)
            financial_year = FinancialYear.objects.get(id=financial_year_id)
        except (
            Organization.DoesNotExist,
            Department.DoesNotExist,
            FinancialYear.DoesNotExist,
        ):
            messages.error(
                request, "Invalid organization, department, or financial year selected."
            )
            return redirect("add_project")

        endorsement = EndorsementForm(
            year=financial_year,
            Project_Title=project_title,
            PSRN=psrn,
            Name_PI=name_pi,
            Designation_PI=designation_pi,
            organization=organization,
            proposal_file=project_proposal,
            sanction_letter=sanction_letter,
            funding_agency_type=funding_agency_type,
            department=department,
            project_type=project_type,
            funding_agency=funding_agency,
            CAPEX=capex,
            OPEX=opex,
            project_duration=project_duration,
            start_date=start_date,
            end_date=end_date,
        )
        endorsement.save()

        co_pis = PrincipleInvestigator.objects.filter(id__in=co_pi_ids)

        if co_pis.exists():
            if len(co_pis) > 0:
                endorsement.Co_PI1 = co_pis[0]
            if len(co_pis) > 1:
                endorsement.Co_PI2 = co_pis[1]
            if len(co_pis) > 2:
                endorsement.Co_PI3 = co_pis[2]
            if len(co_pis) > 3:
                endorsement.Co_PI4 = co_pis[3]
            if len(co_pis) > 4:
                endorsement.Co_PI5 = co_pis[4]
            if len(co_pis) > 5:
                endorsement.Co_PI6 = co_pis[5]

        endorsement.save()

        messages.success(request, "Project added successfully!")
        return redirect("add_project")

    organizations = Organization.objects.all()
    departments = Department.objects.all()
    principal_investigators = PrincipleInvestigator.objects.filter(user=request.user)
    financial_years = FinancialYear.objects.all()

    context = {
        "organizations": organizations,
        "departments": departments,
        "principal_investigators": principal_investigators,
        "financial_years": financial_years,
    }

    return render(request, "management/add_project.html", context)


def bifurcate_budget(request, project_id, installment_id):
    if not request.user.is_authenticated:
        return redirect("login")
    project = get_object_or_404(EndorsementForm, id=project_id)
    installment = get_object_or_404(Installment, id=installment_id)

    budget, created = FinYearBudget.objects.get_or_create(
        project=project, year=installment.installment_year
    )

    if request.method == "POST":

        equipment = request.POST.get("Equipment", 0) or 0
        consumables = request.POST.get("Consumables", 0) or 0
        fellowship = request.POST.get("Fellowship", 0) or 0
        contingency = request.POST.get("Contingency", 0) or 0
        travel = request.POST.get("Travel", 0) or 0
        field_testing = request.POST.get("Field_Testing", 0) or 0
        miscellaneous = request.POST.get("Miscellaneous", 0) or 0

        budget.Equipment = int(equipment)
        budget.Consumables = int(consumables)
        budget.Fellowship = int(fellowship)
        budget.Contingency = int(contingency)
        budget.Travel = int(travel)
        budget.Field_Testing = int(field_testing)
        budget.Miscellaneous = int(miscellaneous)
        budget.save()

        messages.success(request, "Budget bifurcation updated successfully.")
        return redirect("add_installment", project_id=project.id)

    context = {
        "project": project,
        "installment": installment,
        "budget": budget,
    }

    return render(request, "management/bifurcate_budget.html", context)


def export_projects_excel(request):
    if not request.user.is_authenticated:
        return redirect("login")

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects List"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    centered_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    hyperlink_font = Font(color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    main_headers = [
        {"header": "ID", "width": 1, "freeze": True},
        {"header": "Year of Project Sanction", "width": 1, "freeze": True},
        {"header": "Type Of Project", "width": 1, "freeze": True},
        {"header": "Funding Agency Type", "width": 1, "freeze": True},
        {"header": "PI", "width": 1, "freeze": True},
        {"header": "Principal Investigator", "width": 4},
        {"header": "Co. PI-1", "width": 5},
        {"header": "Co. PI-2", "width": 5},
        {"header": "Co. PI-3", "width": 5},
        {"header": "Co. PI-4", "width": 5},
        {"header": "Co. PI-5", "width": 5},
        {"header": "Co. PI-6", "width": 5},
        {"header": "Funding Agency", "width": 1},
        {"header": "Scheme", "width": 1},
        {"header": "Project Name", "width": 1},
        {"header": "Sanctioned-Amount (Rs)", "width": 3},
        {"header": "Project Duration (Months)", "width": 1},
        {"header": "Sanctioned Date", "width": 1},
        {"header": "Project Start Date", "width": 1},
        {"header": "Project End Date", "width": 1},
        {"header": "Total Amount Received (Rs)", "width": 1},
        {"header": "Sanctioned Order", "width": 1},
        {"header": "Project Proposal", "width": 1},
        {"header": "Sanction Letter", "width": 1},
        {"header": "Status", "width": 1},
    ]

    current_column = 1
    for header_info in main_headers:
        cell = ws.cell(row=1, column=current_column)
        cell.value = header_info["header"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = thin_border

        if header_info["width"] > 1:
            ws.merge_cells(
                start_row=1,
                start_column=current_column,
                end_row=1,
                end_column=current_column + header_info["width"] - 1,
            )

        current_column += header_info["width"]

    current_column = 5
    subheaders = ["Name", "PSRN", "Designation", "Campus", "Department"]

    for _ in range(7):
        for subheader in subheaders:
            cell = ws.cell(row=2, column=current_column)
            cell.value = subheader
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = thin_border
            current_column += 1

    amount_start_column = 1
    for header in main_headers[:15]:
        amount_start_column += header["width"]

    for idx, subheader in enumerate(["CAPEX", "RECURRING", "Total"]):
        cell = ws.cell(row=2, column=amount_start_column + idx)
        cell.value = subheader
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = thin_border

    projects = EndorsementForm.objects.all()

    row_num = 3
    for project in projects:
        budget = Budget.objects.filter(project=project).first()
        received_amount = RecievedAmount.objects.filter(project=project).first()

        total_received = 0
        if received_amount:
            total_received = sum(
                [
                    received_amount.Equipment or 0,
                    received_amount.Consumables or 0,
                    received_amount.Fellowship or 0,
                    received_amount.Contingency or 0,
                    received_amount.Travel or 0,
                    received_amount.Field_Testing or 0,
                    received_amount.Miscellaneous or 0,
                ]
            )

        col_num = 1
        basic_data = [
            project.id,
            str(project.year) if project.year else "-",
            project.get_project_type_display(),
            project.get_funding_agency_type_display(),
        ]

        for value in basic_data:
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = centered_alignment
            cell.border = thin_border
            col_num += 1

        pi_data = [
            project.Name_PI or "-",
            project.PSRN or "-",
            project.Designation_PI or "-",
            project.organization.name if project.organization else "-",
            project.department.name if project.department else "-",
        ]

        for value in pi_data:
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = centered_alignment
            cell.border = thin_border
            col_num += 1

        for i in range(1, 7):
            co_pi = getattr(project, f"Co_PI{i}")
            co_pi_data = (
                [
                    co_pi.Name_PI,
                    co_pi.PSRN,
                    co_pi.Designation_PI,
                    co_pi.organization.name if co_pi.organization else "-",
                    co_pi.department.name if co_pi.department else "-",
                ]
                if co_pi
                else ["-"] * 5
            )

            for value in co_pi_data:
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = centered_alignment
                cell.border = thin_border
                col_num += 1

        project_data = [
            project.funding_agency or "-",
            project.funding_scheme or "-",
            project.Project_Title or "-",
        ]

        for value in project_data:
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = centered_alignment
            cell.border = thin_border
            col_num += 1

        for value in [
            project.CAPEX or 0,
            project.OPEX or 0,
            (project.CAPEX or 0) + (project.OPEX or 0),
        ]:
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = centered_alignment
            cell.border = thin_border
            col_num += 1

        remaining_data = [
            project.project_duration or "-",
            (
                budget.sanctiondate.strftime("%d-%m-%Y")
                if budget and budget.sanctiondate
                else "-"
            ),
            project.start_date.strftime("%d-%m-%Y") if project.start_date else "-",
            project.end_date.strftime("%d-%m-%Y") if project.end_date else "-",
            total_received,
            budget.sanctionorder if budget and budget.sanctionorder else "-",
        ]

        for value in remaining_data:
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = centered_alignment
            cell.border = thin_border
            col_num += 1

        for file_field in [project.proposal_file, project.sanction_letter]:
            cell = ws.cell(row=row_num, column=col_num)
            if file_field:
                file_url = request.build_absolute_uri(file_field.url)
                cell.value = "View File"
                cell.hyperlink = file_url
                cell.font = hyperlink_font
            else:
                cell.value = "No file"
            cell.alignment = centered_alignment
            cell.border = thin_border
            col_num += 1

        cell = ws.cell(row=row_num, column=col_num)
        cell.value = project.status.title() if project.status else "-"
        cell.alignment = centered_alignment
        cell.border = thin_border
        col_num += 1

        row_num += 1

    freeze_col = sum(
        header["width"] for header in main_headers if header.get("freeze", False)
    )
    ws.freeze_panes = ws.cell(row=3, column=freeze_col + 1)

    for col_num in range(1, ws.max_column + 1):
        column = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.rows:
        ws.row_dimensions[row[0].row].height = 30

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=Projects_List_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    wb.save(response)
    return response
